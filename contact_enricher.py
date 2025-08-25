"""
Contact Enrichment Module - Enriches existing contact lists with missing information
"""
import csv
import json
import logging
import time
from pathlib import Path
from typing import List, Dict, Optional, Tuple, Any
from dataclasses import dataclass, field, asdict
from datetime import datetime
import openpyxl
from rich.console import Console
from rich.progress import Progress, SpinnerColumn, TextColumn, BarColumn
from rich.table import Table

from perplexity_client import PerplexityClient, ContactInfo

logger = logging.getLogger(__name__)
console = Console()

@dataclass
class EnrichmentRequest:
    """Represents a contact to be enriched"""
    name: str = ""
    company: str = ""
    title: str = ""
    email: str = ""
    phone: str = ""
    address: str = ""
    city: str = ""
    state: str = ""
    country: str = ""
    linkedin: str = ""
    website: str = ""
    notes: str = ""
    original_data: Dict[str, Any] = field(default_factory=dict)
    row_number: int = 0

@dataclass
class EnrichmentResult:
    """Result of enriching a contact"""
    original: EnrichmentRequest
    enriched: Optional[ContactInfo] = None
    status: str = "pending"  # pending, enriched, failed, skipped
    error_message: str = ""
    enrichment_source: str = ""
    confidence: float = 0.0
    timestamp: str = field(default_factory=lambda: datetime.now().isoformat())

class ContactParser:
    """Parses contact lists from various file formats"""
    
    @staticmethod
    def parse_file(file_path: str) -> List[EnrichmentRequest]:
        """Parse a file and return list of contacts to enrich"""
        file_path = Path(file_path)
        
        if not file_path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")
        
        extension = file_path.suffix.lower()
        
        if extension == '.csv':
            return ContactParser._parse_csv(file_path)
        elif extension in ['.xlsx', '.xls']:
            return ContactParser._parse_excel(file_path)
        elif extension == '.json':
            return ContactParser._parse_json(file_path)
        else:
            raise ValueError(f"Unsupported file format: {extension}")
    
    @staticmethod
    def _parse_csv(file_path: Path) -> List[EnrichmentRequest]:
        """Parse CSV file"""
        contacts = []
        
        with open(file_path, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            
            for idx, row in enumerate(reader, 1):
                # Map common field names
                contact = EnrichmentRequest(
                    name=row.get('name', row.get('Name', row.get('full_name', ''))),
                    company=row.get('company', row.get('Company', row.get('organization', ''))),
                    title=row.get('title', row.get('Title', row.get('position', ''))),
                    email=row.get('email', row.get('Email', row.get('email_address', ''))),
                    phone=row.get('phone', row.get('Phone', row.get('phone_number', ''))),
                    address=row.get('address', row.get('Address', row.get('street', ''))),
                    city=row.get('city', row.get('City', '')),
                    state=row.get('state', row.get('State', '')),
                    country=row.get('country', row.get('Country', '')),
                    linkedin=row.get('linkedin', row.get('LinkedIn', row.get('linkedin_url', ''))),
                    website=row.get('website', row.get('Website', row.get('company_website', ''))),
                    notes=row.get('notes', row.get('Notes', '')),
                    original_data=dict(row),
                    row_number=idx
                )
                contacts.append(contact)
        
        return contacts
    
    @staticmethod
    def _parse_excel(file_path: Path) -> List[EnrichmentRequest]:
        """Parse Excel file"""
        contacts = []
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        
        # Get headers from first row
        headers = []
        for cell in sheet[1]:
            headers.append(cell.value if cell.value else '')
        
        # Parse data rows
        for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 1):
            row_dict = {headers[i]: row[i] for i in range(len(headers)) if i < len(row)}
            
            contact = EnrichmentRequest(
                name=row_dict.get('name', row_dict.get('Name', '')),
                company=row_dict.get('company', row_dict.get('Company', '')),
                title=row_dict.get('title', row_dict.get('Title', '')),
                email=row_dict.get('email', row_dict.get('Email', '')),
                phone=row_dict.get('phone', row_dict.get('Phone', '')),
                address=row_dict.get('address', row_dict.get('Address', '')),
                city=row_dict.get('city', row_dict.get('City', '')),
                state=row_dict.get('state', row_dict.get('State', '')),
                country=row_dict.get('country', row_dict.get('Country', '')),
                linkedin=row_dict.get('linkedin', row_dict.get('LinkedIn', '')),
                website=row_dict.get('website', row_dict.get('Website', '')),
                notes=row_dict.get('notes', row_dict.get('Notes', '')),
                original_data=row_dict,
                row_number=idx
            )
            contacts.append(contact)
        
        return contacts
    
    @staticmethod
    def _parse_json(file_path: Path) -> List[EnrichmentRequest]:
        """Parse JSON file"""
        contacts = []
        
        with open(file_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        # Handle both array and object with contacts array
        if isinstance(data, dict):
            data = data.get('contacts', data.get('data', []))
        
        for idx, item in enumerate(data, 1):
            contact = EnrichmentRequest(
                name=item.get('name', ''),
                company=item.get('company', ''),
                title=item.get('title', ''),
                email=item.get('email', ''),
                phone=item.get('phone', ''),
                address=item.get('address', ''),
                city=item.get('city', ''),
                state=item.get('state', ''),
                country=item.get('country', ''),
                linkedin=item.get('linkedin', ''),
                website=item.get('website', ''),
                notes=item.get('notes', ''),
                original_data=item,
                row_number=idx
            )
            contacts.append(contact)
        
        return contacts

class ContactEnricher:
    """Enriches contact information using Perplexity API"""
    
    def __init__(self, perplexity_client: PerplexityClient, 
                 rate_limit_delay: float = 1.0,
                 batch_size: int = 10):
        """Initialize enricher"""
        self.client = perplexity_client
        self.rate_limit_delay = rate_limit_delay
        self.batch_size = batch_size
        self.results: List[EnrichmentResult] = []
        self.state_file = Path("enrichment_state.json")
    
    def enrich_contacts(self, contacts: List[EnrichmentRequest], 
                        resume: bool = False) -> List[EnrichmentResult]:
        """Enrich a list of contacts"""
        
        # Load previous state if resuming
        start_idx = 0
        if resume and self.state_file.exists():
            state = self._load_state()
            self.results = state.get('results', [])
            start_idx = len(self.results)
            console.print(f"[yellow]Resuming from contact {start_idx + 1}[/yellow]")
        
        total = len(contacts)
        
        with Progress(
            SpinnerColumn(),
            TextColumn("[progress.description]{task.description}"),
            BarColumn(),
            TextColumn("[progress.percentage]{task.percentage:>3.0f}%"),
            console=console
        ) as progress:
            
            task = progress.add_task(
                f"Enriching {total} contacts...", 
                total=total
            )
            
            for idx, contact in enumerate(contacts[start_idx:], start_idx):
                progress.update(task, advance=1, 
                              description=f"Enriching {contact.name or contact.company}...")
                
                result = self._enrich_single_contact(contact)
                self.results.append(result)
                
                # Save state periodically
                if (idx + 1) % 5 == 0:
                    self._save_state()
                
                # Rate limiting
                time.sleep(self.rate_limit_delay)
        
        # Final save
        self._save_state()
        
        # Print summary
        self._print_summary()
        
        return self.results
    
    def _enrich_single_contact(self, contact: EnrichmentRequest) -> EnrichmentResult:
        """Enrich a single contact"""
        
        # Skip if we already have both email and phone
        if contact.email and contact.phone:
            return EnrichmentResult(
                original=contact,
                status="skipped",
                error_message="Already has email and phone"
            )
        
        # Build enrichment query
        query = self._build_enrichment_query(contact)
        
        if not query:
            return EnrichmentResult(
                original=contact,
                status="failed",
                error_message="Insufficient information to enrich"
            )
        
        try:
            # Search for enriched information
            enriched_contacts = self.client.search_contact(
                query=query,
                additional_context=self._build_context(contact)
            )
            
            if enriched_contacts:
                # Find best match
                best_match = self._find_best_match(contact, enriched_contacts)
                
                if best_match:
                    return EnrichmentResult(
                        original=contact,
                        enriched=best_match,
                        status="enriched",
                        enrichment_source="Perplexity AI",
                        confidence=best_match.confidence_score
                    )
            
            return EnrichmentResult(
                original=contact,
                status="failed",
                error_message="No enrichment data found"
            )
            
        except Exception as e:
            logger.error(f"Error enriching contact {contact.name}: {str(e)}")
            return EnrichmentResult(
                original=contact,
                status="failed",
                error_message=str(e)
            )
    
    def _build_enrichment_query(self, contact: EnrichmentRequest) -> str:
        """Build query for enrichment"""
        parts = []
        
        # Name and company are most important
        if contact.name and contact.company:
            parts.append(f"{contact.name} {contact.company}")
        elif contact.company:
            parts.append(f"{contact.company} owner CEO contact information")
        elif contact.name:
            parts.append(contact.name)
        else:
            return ""
        
        # Add title if available
        if contact.title:
            parts.append(contact.title)
        
        # Add location if available
        if contact.city and contact.state:
            parts.append(f"{contact.city} {contact.state}")
        elif contact.state:
            parts.append(contact.state)
        
        # Request missing information
        missing = []
        if not contact.email:
            missing.append("email")
        if not contact.phone:
            missing.append("phone number")
        
        if missing:
            parts.append(f"find {' and '.join(missing)}")
        
        return " ".join(parts)
    
    def _build_context(self, contact: EnrichmentRequest) -> str:
        """Build additional context for the search"""
        context_parts = []
        
        if contact.website:
            context_parts.append(f"Company website: {contact.website}")
        
        if contact.linkedin:
            context_parts.append(f"LinkedIn: {contact.linkedin}")
        
        if contact.address:
            context_parts.append(f"Address: {contact.address}")
        
        if contact.notes:
            context_parts.append(f"Notes: {contact.notes}")
        
        # Specify what we're looking for
        missing = []
        if not contact.email:
            missing.append("direct email address")
        if not contact.phone:
            missing.append("direct phone number")
        
        if missing:
            context_parts.append(f"IMPORTANT: Find the {' and '.join(missing)} for this specific person/company")
        
        return "\n".join(context_parts)
    
    def _find_best_match(self, original: EnrichmentRequest, 
                        candidates: List[ContactInfo]) -> Optional[ContactInfo]:
        """Find the best matching contact from candidates"""
        
        best_match = None
        best_score = 0
        
        for candidate in candidates:
            score = 0
            
            # Name matching
            if original.name and candidate.name:
                if original.name.lower() in candidate.name.lower() or \
                   candidate.name.lower() in original.name.lower():
                    score += 3
            
            # Company matching
            if original.company and candidate.company:
                if original.company.lower() in candidate.company.lower() or \
                   candidate.company.lower() in original.company.lower():
                    score += 3
            
            # Has the missing information we need
            if not original.email and candidate.primary_email:
                score += 2
            if not original.phone and candidate.primary_phone:
                score += 2
            
            # Confidence score
            score += candidate.confidence_score
            
            if score > best_score:
                best_score = score
                best_match = candidate
        
        return best_match
    
    def _save_state(self):
        """Save current enrichment state"""
        state = {
            'timestamp': datetime.now().isoformat(),
            'results': [
                {
                    'original': asdict(r.original),
                    'status': r.status,
                    'error_message': r.error_message,
                    'confidence': r.confidence,
                    'enriched': asdict(r.enriched) if r.enriched else None
                }
                for r in self.results
            ]
        }
        
        with open(self.state_file, 'w') as f:
            json.dump(state, f, indent=2)
    
    def _load_state(self) -> Dict:
        """Load previous enrichment state"""
        with open(self.state_file, 'r') as f:
            return json.load(f)
    
    def _print_summary(self):
        """Print enrichment summary"""
        total = len(self.results)
        enriched = sum(1 for r in self.results if r.status == "enriched")
        failed = sum(1 for r in self.results if r.status == "failed")
        skipped = sum(1 for r in self.results if r.status == "skipped")
        
        table = Table(title="Enrichment Summary")
        table.add_column("Status", style="cyan")
        table.add_column("Count", style="magenta")
        table.add_column("Percentage", style="green")
        
        table.add_row("Enriched", str(enriched), f"{enriched/total*100:.1f}%")
        table.add_row("Failed", str(failed), f"{failed/total*100:.1f}%")
        table.add_row("Skipped", str(skipped), f"{skipped/total*100:.1f}%")
        table.add_row("Total", str(total), "100%")
        
        console.print(table)

class EnrichmentExporter:
    """Export enriched contact data"""
    
    def __init__(self, output_dir: str = "enriched_output"):
        """Initialize exporter"""
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(exist_ok=True)
    
    def export_results(self, results: List[EnrichmentResult], 
                      format: str = "csv") -> str:
        """Export enrichment results"""
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        if format == "csv":
            return self._export_csv(results, timestamp)
        elif format == "excel":
            return self._export_excel(results, timestamp)
        elif format == "json":
            return self._export_json(results, timestamp)
        else:
            raise ValueError(f"Unsupported format: {format}")
    
    def _export_csv(self, results: List[EnrichmentResult], timestamp: str) -> str:
        """Export to CSV with original and enriched data side by side"""
        
        filename = f"enriched_contacts_{timestamp}.csv"
        filepath = self.output_dir / filename
        
        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            fieldnames = [
                'Row', 'Status',
                'Original_Name', 'Original_Company', 'Original_Title',
                'Original_Email', 'Original_Phone',
                'Original_City', 'Original_State',
                'Enriched_Name', 'Enriched_Company',
                'Enriched_Email', 'Enriched_Alt_Emails',
                'Enriched_Phone', 'Enriched_Alt_Phones',
                'Confidence', 'Sources', 'Notes', 'Error'
            ]
            
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            
            for result in results:
                row = {
                    'Row': result.original.row_number,
                    'Status': result.status,
                    'Original_Name': result.original.name,
                    'Original_Company': result.original.company,
                    'Original_Title': result.original.title,
                    'Original_Email': result.original.email,
                    'Original_Phone': result.original.phone,
                    'Original_City': result.original.city,
                    'Original_State': result.original.state,
                    'Confidence': result.confidence,
                    'Error': result.error_message
                }
                
                if result.enriched:
                    row.update({
                        'Enriched_Name': result.enriched.name,
                        'Enriched_Company': result.enriched.company,
                        'Enriched_Email': result.enriched.primary_email,
                        'Enriched_Alt_Emails': ', '.join(result.enriched.alternate_emails),
                        'Enriched_Phone': result.enriched.primary_phone,
                        'Enriched_Alt_Phones': ', '.join(result.enriched.alternate_phones),
                        'Sources': '; '.join([s['url'] for s in result.enriched.sources[:3]]),
                        'Notes': result.enriched.notes
                    })
                
                writer.writerow(row)
        
        return str(filepath)
    
    def _export_excel(self, results: List[EnrichmentResult], timestamp: str) -> str:
        """Export to Excel with multiple sheets"""
        
        filename = f"enriched_contacts_{timestamp}.xlsx"
        filepath = self.output_dir / filename
        
        workbook = openpyxl.Workbook()
        
        # Summary sheet
        summary_sheet = workbook.active
        summary_sheet.title = "Summary"
        
        # Enriched contacts sheet
        enriched_sheet = workbook.create_sheet("Enriched")
        self._write_enriched_sheet(enriched_sheet, 
                                   [r for r in results if r.status == "enriched"])
        
        # Failed contacts sheet
        failed_sheet = workbook.create_sheet("Failed")
        self._write_failed_sheet(failed_sheet,
                                [r for r in results if r.status == "failed"])
        
        # Write summary
        self._write_summary_sheet(summary_sheet, results)
        
        workbook.save(filepath)
        return str(filepath)
    
    def _write_enriched_sheet(self, sheet, results):
        """Write enriched contacts to sheet"""
        headers = [
            'Name', 'Company', 'Title', 'Email', 'Phone',
            'Alt Emails', 'Alt Phones', 'City', 'State',
            'Confidence', 'Sources'
        ]
        
        # Write headers
        for col, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col, value=header)
        
        # Write data
        for row, result in enumerate(results, 2):
            sheet.cell(row=row, column=1, value=result.enriched.name)
            sheet.cell(row=row, column=2, value=result.enriched.company)
            sheet.cell(row=row, column=3, value=result.original.title)
            sheet.cell(row=row, column=4, value=result.enriched.primary_email)
            sheet.cell(row=row, column=5, value=result.enriched.primary_phone)
            sheet.cell(row=row, column=6, value=', '.join(result.enriched.alternate_emails))
            sheet.cell(row=row, column=7, value=', '.join(result.enriched.alternate_phones))
            sheet.cell(row=row, column=8, value=result.original.city)
            sheet.cell(row=row, column=9, value=result.original.state)
            sheet.cell(row=row, column=10, value=f"{result.confidence:.2f}")
            sheet.cell(row=row, column=11, value='; '.join([s['url'] for s in result.enriched.sources[:2]]))
    
    def _write_failed_sheet(self, sheet, results):
        """Write failed contacts to sheet"""
        headers = ['Name', 'Company', 'Error Message']
        
        for col, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col, value=header)
        
        for row, result in enumerate(results, 2):
            sheet.cell(row=row, column=1, value=result.original.name)
            sheet.cell(row=row, column=2, value=result.original.company)
            sheet.cell(row=row, column=3, value=result.error_message)
    
    def _write_summary_sheet(self, sheet, results):
        """Write summary statistics"""
        total = len(results)
        enriched = sum(1 for r in results if r.status == "enriched")
        failed = sum(1 for r in results if r.status == "failed")
        skipped = sum(1 for r in results if r.status == "skipped")
        
        sheet.cell(row=1, column=1, value="Enrichment Summary")
        sheet.cell(row=3, column=1, value="Total Contacts:")
        sheet.cell(row=3, column=2, value=total)
        sheet.cell(row=4, column=1, value="Successfully Enriched:")
        sheet.cell(row=4, column=2, value=enriched)
        sheet.cell(row=5, column=1, value="Failed:")
        sheet.cell(row=5, column=2, value=failed)
        sheet.cell(row=6, column=1, value="Skipped:")
        sheet.cell(row=6, column=2, value=skipped)
        sheet.cell(row=8, column=1, value="Success Rate:")
        sheet.cell(row=8, column=2, value=f"{enriched/total*100:.1f}%" if total > 0 else "0%")
    
    def _export_json(self, results: List[EnrichmentResult], timestamp: str) -> str:
        """Export to JSON format"""
        
        filename = f"enriched_contacts_{timestamp}.json"
        filepath = self.output_dir / filename
        
        data = {
            'timestamp': timestamp,
            'summary': {
                'total': len(results),
                'enriched': sum(1 for r in results if r.status == "enriched"),
                'failed': sum(1 for r in results if r.status == "failed"),
                'skipped': sum(1 for r in results if r.status == "skipped")
            },
            'results': []
        }
        
        for result in results:
            item = {
                'status': result.status,
                'original': asdict(result.original),
                'enriched': asdict(result.enriched) if result.enriched else None,
                'confidence': result.confidence,
                'error': result.error_message,
                'timestamp': result.timestamp
            }
            data['results'].append(item)
        
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2)
        
        return str(filepath)