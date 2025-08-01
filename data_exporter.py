"""
Data export functionality for contact information
"""
import csv
import json
from pathlib import Path
from typing import List, Dict
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from perplexity_client import ContactInfo

class DataExporter:
    """Export contact data to various formats"""
    
    def __init__(self, output_dir: str = "output"):
        """Initialize exporter with output directory"""
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(exist_ok=True)
    
    def export_to_csv(self, contacts: List[ContactInfo], filename: str = None) -> str:
        """Export contacts to CSV format"""
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"contacts_{timestamp}.csv"
        
        filepath = self.output_dir / filename
        
        with open(filepath, 'w', newline='', encoding='utf-8') as csvfile:
            fieldnames = [
                'Name', 'Company', 'Primary Email', 'Alternate Emails',
                'Primary Phone', 'Alternate Phones', 'Sources', 
                'Confidence', 'Email Status', 'Phone Status', 'Notes', 'Date Found'
            ]
            
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            writer.writeheader()
            
            for contact in contacts:
                # Format sources
                sources_str = self._format_sources(contact.sources)
                
                # Format alternate contacts
                alt_emails = ', '.join(contact.alternate_emails)
                alt_phones = ', '.join(contact.alternate_phones)
                
                writer.writerow({
                    'Name': contact.name,
                    'Company': contact.company,
                    'Primary Email': contact.primary_email,
                    'Alternate Emails': alt_emails,
                    'Primary Phone': contact.primary_phone,
                    'Alternate Phones': alt_phones,
                    'Sources': sources_str,
                    'Confidence': f"{contact.confidence_score:.2f}",
                    'Email Status': contact.verification_status.get('primary_email', 'unverified'),
                    'Phone Status': contact.verification_status.get('primary_phone', 'unverified'),
                    'Notes': contact.notes,
                    'Date Found': contact.date_found
                })
        
        print(f"Exported {len(contacts)} contacts to {filepath}")
        return str(filepath)
    
    def export_to_apollo_csv(self, contacts: List[ContactInfo], filename: str = None) -> str:
        """Export contacts in Apollo-compatible CSV format"""
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"apollo_contacts_{timestamp}.csv"
        
        filepath = self.output_dir / filename
        
        with open(filepath, 'w', newline='', encoding='utf-8') as csvfile:
            # Apollo-specific fields
            fieldnames = [
                'First Name', 'Last Name', 'Email', 'Company', 
                'Phone Number', 'Mobile Phone', 'Title', 'LinkedIn URL'
            ]
            
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            writer.writeheader()
            
            for contact in contacts:
                # Split name into first and last
                name_parts = contact.name.split(' ', 1)
                first_name = name_parts[0] if name_parts else ''
                last_name = name_parts[1] if len(name_parts) > 1 else ''
                
                # Find LinkedIn URL in sources
                linkedin_url = self._find_linkedin_url(contact.sources)
                
                # Primary row
                writer.writerow({
                    'First Name': first_name,
                    'Last Name': last_name,
                    'Email': contact.primary_email,
                    'Company': contact.company,
                    'Phone Number': contact.primary_phone,
                    'Mobile Phone': contact.alternate_phones[0] if contact.alternate_phones else '',
                    'Title': '',  # Would need to extract from notes or response
                    'LinkedIn URL': linkedin_url
                })
                
                # Additional rows for alternate emails
                for alt_email in contact.alternate_emails:
                    writer.writerow({
                        'First Name': first_name,
                        'Last Name': last_name,
                        'Email': alt_email,
                        'Company': contact.company,
                        'Phone Number': contact.primary_phone,
                        'Mobile Phone': '',
                        'Title': '',
                        'LinkedIn URL': linkedin_url
                    })
        
        print(f"Exported {len(contacts)} contacts to Apollo format: {filepath}")
        return str(filepath)
    
    def export_to_json(self, contacts: List[ContactInfo], filename: str = None) -> str:
        """Export contacts to JSON format with full details"""
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"contacts_{timestamp}.json"
        
        filepath = self.output_dir / filename
        
        # Convert ContactInfo objects to dictionaries
        contacts_data = []
        for contact in contacts:
            contact_dict = {
                'name': contact.name,
                'company': contact.company,
                'primary_email': contact.primary_email,
                'alternate_emails': contact.alternate_emails,
                'primary_phone': contact.primary_phone,
                'alternate_phones': contact.alternate_phones,
                'sources': contact.sources,
                'confidence_score': contact.confidence_score,
                'verification_status': contact.verification_status,
                'notes': contact.notes,
                'date_found': contact.date_found,
                'raw_response': contact.raw_response
            }
            contacts_data.append(contact_dict)
        
        with open(filepath, 'w', encoding='utf-8') as jsonfile:
            json.dump(contacts_data, jsonfile, indent=2, ensure_ascii=False)
        
        print(f"Exported {len(contacts)} contacts to {filepath}")
        return str(filepath)
    
    def export_summary(self, contacts: List[ContactInfo]) -> Dict[str, any]:
        """Generate a summary of the export"""
        total_contacts = len(contacts)
        verified_emails = sum(1 for c in contacts if c.verification_status.get('primary_email') == 'valid')
        verified_phones = sum(1 for c in contacts if c.verification_status.get('primary_phone') == 'valid')
        
        total_emails = sum(1 + len(c.alternate_emails) for c in contacts if c.primary_email)
        total_phones = sum(1 + len(c.alternate_phones) for c in contacts if c.primary_phone)
        
        companies = set(c.company for c in contacts if c.company)
        
        summary = {
            'total_contacts': total_contacts,
            'contacts_with_email': sum(1 for c in contacts if c.primary_email),
            'contacts_with_phone': sum(1 for c in contacts if c.primary_phone),
            'verified_emails': verified_emails,
            'verified_phones': verified_phones,
            'total_email_addresses': total_emails,
            'total_phone_numbers': total_phones,
            'unique_companies': len(companies),
            'average_confidence': sum(c.confidence_score for c in contacts) / total_contacts if total_contacts > 0 else 0
        }
        
        return summary
    
    def _format_sources(self, sources: List[Dict[str, str]]) -> str:
        """Format sources for CSV output"""
        if not sources:
            return ""
        
        formatted = []
        for source in sources:
            url = source.get('url', '')
            title = source.get('title', 'Source')
            relevance = source.get('relevance', '')
            
            if relevance:
                formatted.append(f"{url} ({title} - {relevance})")
            else:
                formatted.append(f"{url} ({title})")
        
        return "; ".join(formatted)
    
    def _find_linkedin_url(self, sources: List[Dict[str, str]]) -> str:
        """Extract LinkedIn URL from sources"""
        for source in sources:
            url = source.get('url', '')
            if 'linkedin.com' in url:
                return url
        return ""
    
    def export_to_excel(self, contacts: List[ContactInfo], filename: str = None) -> str:
        """Export contacts to Excel format with formatting"""
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"contacts_{timestamp}.xlsx"
        
        filepath = self.output_dir / filename
        
        # Create workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Contacts"
        
        # Define headers
        headers = [
            'Name', 'Company', 'Primary Email', 'Alternate Emails',
            'Primary Phone', 'Alternate Phones', 'Sources', 
            'Source URLs', 'Confidence', 'Email Status', 'Phone Status', 
            'Notes', 'Date Found'
        ]
        
        # Style headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Add data
        for row, contact in enumerate(contacts, 2):
            # Format sources for readability
            sources_titles = []
            sources_urls = []
            for source in contact.sources:
                sources_titles.append(source.get('title', 'Source'))
                sources_urls.append(source.get('url', ''))
            
            ws.cell(row=row, column=1, value=contact.name)
            ws.cell(row=row, column=2, value=contact.company)
            ws.cell(row=row, column=3, value=contact.primary_email)
            ws.cell(row=row, column=4, value=', '.join(contact.alternate_emails))
            ws.cell(row=row, column=5, value=contact.primary_phone)
            ws.cell(row=row, column=6, value=', '.join(contact.alternate_phones))
            ws.cell(row=row, column=7, value='\n'.join(sources_titles))
            ws.cell(row=row, column=8, value='\n'.join(sources_urls))
            ws.cell(row=row, column=9, value=f"{contact.confidence_score:.2f}")
            ws.cell(row=row, column=10, value=contact.verification_status.get('primary_email', 'unverified'))
            ws.cell(row=row, column=11, value=contact.verification_status.get('primary_phone', 'unverified'))
            ws.cell(row=row, column=12, value=contact.notes)
            ws.cell(row=row, column=13, value=contact.date_found)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save workbook
        wb.save(filepath)
        print(f"Exported {len(contacts)} contacts to Excel: {filepath}")
        return str(filepath)
    
    def export_to_txt(self, contacts: List[ContactInfo], filename: str = None) -> str:
        """Export contacts to formatted text file"""
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"contacts_{timestamp}.txt"
        
        filepath = self.output_dir / filename
        
        with open(filepath, 'w', encoding='utf-8') as txtfile:
            txtfile.write("CONTACT FINDER RESULTS\n")
            txtfile.write("=" * 80 + "\n")
            txtfile.write(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            txtfile.write(f"Total Contacts: {len(contacts)}\n")
            txtfile.write("=" * 80 + "\n\n")
            
            for i, contact in enumerate(contacts, 1):
                txtfile.write(f"CONTACT #{i}\n")
                txtfile.write("-" * 40 + "\n")
                txtfile.write(f"Name: {contact.name}\n")
                txtfile.write(f"Company: {contact.company}\n")
                txtfile.write(f"Confidence Score: {contact.confidence_score:.0%}\n")
                txtfile.write("\n")
                
                # Email information
                txtfile.write("EMAIL INFORMATION:\n")
                if contact.primary_email:
                    txtfile.write(f"  Primary: {contact.primary_email}")
                    if contact.verification_status.get('primary_email'):
                        txtfile.write(f" [{contact.verification_status.get('primary_email')}]")
                    txtfile.write("\n")
                if contact.alternate_emails:
                    txtfile.write(f"  Alternates: {', '.join(contact.alternate_emails)}\n")
                else:
                    txtfile.write("  No email found\n")
                txtfile.write("\n")
                
                # Phone information
                txtfile.write("PHONE INFORMATION:\n")
                if contact.primary_phone:
                    txtfile.write(f"  Primary: {contact.primary_phone}")
                    if contact.verification_status.get('primary_phone'):
                        txtfile.write(f" [{contact.verification_status.get('primary_phone')}]")
                    txtfile.write("\n")
                if contact.alternate_phones:
                    txtfile.write(f"  Alternates: {', '.join(contact.alternate_phones)}\n")
                else:
                    txtfile.write("  No phone found\n")
                txtfile.write("\n")
                
                # Sources - PROMINENT DISPLAY
                txtfile.write("SOURCES (Verify these yourself):\n")
                if contact.sources:
                    for j, source in enumerate(contact.sources, 1):
                        txtfile.write(f"  {j}. {source.get('title', 'Source')}\n")
                        txtfile.write(f"     URL: {source.get('url', 'N/A')}\n")
                        if source.get('relevance'):
                            txtfile.write(f"     Relevance: {source.get('relevance')}\n")
                else:
                    txtfile.write("  No sources available\n")
                
                if contact.notes:
                    txtfile.write(f"\nNotes: {contact.notes}\n")
                
                txtfile.write("\n" + "=" * 80 + "\n\n")
        
        print(f"Exported {len(contacts)} contacts to text file: {filepath}")
        return str(filepath)
    
    def print_summary(self, contacts: List[ContactInfo]):
        """Print a summary of the results"""
        summary = self.export_summary(contacts)
        
        print("\n" + "="*60)
        print("CONTACT FINDER SUMMARY")
        print("="*60)
        print(f"Total contacts found: {summary['total_contacts']}")
        print(f"Contacts with email: {summary['contacts_with_email']}")
        print(f"Contacts with phone: {summary['contacts_with_phone']}")
        print(f"Verified emails: {summary['verified_emails']}")
        print(f"Verified phones: {summary['verified_phones']}")
        print(f"Total email addresses: {summary['total_email_addresses']}")
        print(f"Total phone numbers: {summary['total_phone_numbers']}")
        print(f"Unique companies: {summary['unique_companies']}")
        print(f"Average confidence: {summary['average_confidence']:.2%}")
        print("="*60 + "\n")